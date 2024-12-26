# src/data/excel_storage.py
import os
from typing import List

import pandas as pd
from openpyxl.reader.excel import load_workbook

from src.core.base_types import Numbers
from .storage_interface import StorageInterface
from ..logging_config import logger


class ExcelStorage(StorageInterface):
    def __init__(self, file_path: str = "xlsx/carrinhos.xlsx"):
        self.file_path = file_path
        self.temp_sheet = '_temp'

    def create_storage(self) -> bool:
        """Cria o arquivo Excel se não existir."""
        try:
            if not os.path.exists(self.file_path):
                df = pd.DataFrame()
                df.to_excel(self.file_path, sheet_name=self.temp_sheet, index=False)
            return True
        except Exception as e:
            logger.error(f"Erro ao criar armazenamento: {e}")
            return False

    def list_carts(self) -> List[str]:
        """Lista todos os carrinhos (sheets) disponíveis."""
        try:
            if not os.path.exists(self.file_path):
                self.create_storage()
                return []

            xls = pd.ExcelFile(self.file_path)
            return [sheet for sheet in xls.sheet_names if sheet != self.temp_sheet]
        except Exception as e:
            logger.error(f"Erro ao listar carrinhos: {e}")
            return []

    def add_cart(self, name: str, tickets: List[Numbers]) -> bool:
        """Adiciona ou atualiza um carrinho com seus jogos."""
        try:
            # Cria DataFrame com os novos jogos
            df_new = pd.DataFrame(tickets, columns=[f'Número {i + 1}' for i in range(6)])

            # Se o carrinho já existe, concatena com jogos existentes
            if self.cart_exists(name):
                df_existing = pd.read_excel(self.file_path, sheet_name=name)
                df_final = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_final = df_new

            # Salva todos os carrinhos
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Salva o carrinho atualizado
                df_final.to_excel(writer, sheet_name=name, index=False)

                # Salva os outros carrinhos sem alteração
                cart_names = self.list_carts()
                for cart in cart_names:
                    if cart != name:
                        df = pd.read_excel(self.file_path, sheet_name=cart)
                        df.to_excel(writer, sheet_name=cart, index=False)

                # Adiciona planilha temporária se não existir
                if self.temp_sheet not in writer.book.sheetnames:
                    pd.DataFrame().to_excel(writer, sheet_name=self.temp_sheet, index=False)

            return True
        except Exception as e:
            logger.error(f"Erro ao adicionar carrinho: {e}")
            return False

    def get_cart_tickets(self, name: str) -> List[Numbers]:
        """Obtém os jogos de um carrinho específico."""
        try:
            if not self.cart_exists(name):
                return []

            df = pd.read_excel(self.file_path, sheet_name=name)
            return df.values.tolist()
        except Exception as e:
            logger.error(f"Erro ao obter cartões: {e}")
            return []

    def delete_cart(self, name: str) -> bool:
        """Remove uma aba (carrinho) específica do arquivo Excel."""
        try:
            if not self.cart_exists(name):
                return False

            # Carrega o arquivo existente
            workbook = load_workbook(self.file_path)

            # Verifica se a aba existe e remove
            if name in workbook.sheetnames:
                workbook.remove(workbook[name])
                workbook.save(self.file_path)
                return True

            return False
        except Exception as e:
            logger.error(f"Erro ao excluir carrinho: {e}")
            return False

    def cart_exists(self, name: str) -> bool:
        """Verifica se um carrinho existe."""
        return name in self.list_carts()
